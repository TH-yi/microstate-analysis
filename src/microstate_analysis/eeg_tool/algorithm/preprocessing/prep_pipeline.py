#!/usr/bin/env python
# -*- coding: utf-8 -*-
# EEG signals respond differently to idea generation, idea evolution, and evaluation in a loosely controlled creativity experiment
# Time    : 2019-10-10
# Author  : Wenjun Jia
# File    : pre_pipeline.py


import time
import numpy as np
import codecs, json
import mne
from scipy import stats
import matlab
import matlab.engine
from pyprep.noisy import Noisydata
from eeg_tool.model.raw_data import RawData
from collections import OrderedDict, Counter
from openpyxl import Workbook, load_workbook
from eeg_tool.utilis import read_subject_info
import os
from multiprocessing import Pool
import warnings
from sklearn.decomposition import FastICA
from scipy.io import savemat

class PreProcessPipeline:
    # engine = None
    engine = matlab.engine.start_matlab()
    def __init__(self, raw, tasks, trial_info, tasks_cleaned=None):
        self.raw = raw
        self.tasks = tasks
        self.trial_info = trial_info
        self.trial_name = []
        self.trial_duration = []
        self.tasks_merged = None
        self.tasks_cleaned = tasks_cleaned
        self.epochs_cleaned = None
        self.global_bads = None
        self.global_good_index = None
        self.global_good_name = None
        self.n_ch = None
        self.n_t = None
        self.fast_ica_convergence = None



    def concatenate_tasks(self):
        flag = True
        duration = []
        onset = [0.]
        for key, value in self.tasks.items():
            n_time = value.shape[1] / self.raw.info['sfreq']
            duration.append(n_time)
            self.trial_name.append(key)
            self.trial_duration.append(0 if len(self.trial_duration) == 0 else self.trial_duration[-1])
            self.trial_duration.append(self.trial_duration[-1] + value.shape[1])
            if flag:
                self.tasks_merged = value
                flag = False
            else:
                onset.append(np.sum(duration))
                self.tasks_merged = np.concatenate((self.tasks_merged, value), axis=1)

        self.tasks_merged = mne.io.RawArray(self.tasks_merged, self.raw.info)
        self.tasks_merged = self.tasks_merged.set_annotations(mne.Annotations(onset, duration, self.trial_name), False)
        self.n_t = self.tasks_merged.n_times
        self.n_ch = len(self.tasks_merged.ch_names)

    def merge_task(self, task_name_combined):
        res = {}
        for task in task_name_combined:
            a_task = task.split("#")
            if len(a_task) > 1:
                temp = self.tasks[a_task[0]]
                for i in range(1, len(a_task)):
                    temp = np.concatenate((temp, self.tasks[a_task[i]]), axis=1)
            else:
                temp = self.tasks[task]
            res[task] = temp
        return res


    def filter(self, low_frequency=1., high_frequency=240.):
        self.tasks_cleaned = self.tasks_merged.copy().filter(low_frequency, high_frequency)

    def remove_line_noise(self):
        self.tasks_cleaned = self.tasks_cleaned.notch_filter(60, filter_length='auto', phase='zero')

    @staticmethod
    def bad_channel(data):
        nd = Noisydata(data)
        nd.find_all_bads()
        bads = nd.get_bads(verbose=False)
        return bads

    def remove_bad_channel(self, thread=1, threshold=0.1):
        pool = Pool(thread)
        res = []
        bads_list = []
        bads = []
        threshold = threshold * len(self.trial_duration) / 2
        for i in range(0, len(self.trial_duration), 2):
            start_index = self.trial_duration[i]
            end_index = self.trial_duration[i + 1]
            data_obj = mne.io.RawArray(filtered_data._data[:, start_index:end_index], self.raw.info)
            res.append(pool.apply_async(PreProcessPipeline.bad_channel, (data_obj,)))
        pool.close()
        pool.join()
        for temp in res:
            bads_list.extend(temp.get())
        bads_set = Counter(bads_list)
        for key, value in bads_set.items():
            if value > threshold:
                bads.append(key)
        filtered_data.info['bads'] = bads
        self.global_bads = bads
        self.global_good_index = np.asarray([i for i in range(self.n_ch) if filtered_data.ch_names[i] not in bads])
        self.global_good_name = [filtered_data.ch_names[i] for i in range(self.n_ch) if filtered_data.ch_names[i] not in bads]
        del pool
        print(bads_set)
        print(bads)
        print(self.global_good_index)

    # def remove_bad_channel(self, thread=5, threshold=0.1):
    #     pool = Pool(thread)
    #     res = []
    #     bads_list = []
    #     bads = []
    #     threshold = threshold * len(self.trial_duration) / 2
    #     for i in range(0, len(self.trial_duration), 2):
    #         start_index = self.trial_duration[i]
    #         end_index = self.trial_duration[i + 1]
    #         data_obj = mne.io.RawArray(self.tasks_cleaned._data[:, start_index:end_index], self.raw.info)
    #         res.append(pool.apply_async(PreProcessPipeline.bad_channel, (data_obj,)))
    #     pool.close()
    #     pool.join()
    #     for temp in res:
    #         bads_list.extend(temp.get())
    #     bads_set = Counter(bads_list)
    #     for key, value in bads_set.items():
    #         if value > threshold:
    #             bads.append(key)
    #     self.tasks_cleaned.info['bads'] = bads
    #     self.global_bads = bads
    #     self.global_good_index = np.asarray([i for i in range(self.n_ch) if self.tasks_cleaned.ch_names[i] not in bads])
    #     self.global_good_name = [self.tasks_cleaned.ch_names[i] for i in range(self.n_ch) if self.tasks_cleaned.ch_names[i] not in bads]
    #     del pool
    #     print(bads_set)
    #     print(bads)
        


    def remove_artifact_wica(self, wave_name='coif5', level=5, multipier=1, fast_ica_iter=3, tol=0.025):
        with warnings.catch_warnings():
            warnings.filterwarnings("error")
            loop_break = False
            while 1:
                for i in range(fast_ica_iter):
                    try:
                        ica = FastICA(max_iter=200, whiten=True, tol=tol)
                        sources = ica.fit_transform(self.tasks_cleaned.get_data(picks=self.global_good_name).T).T
                        loop_break = True
                        break
                    except Warning:
                        print('FastICA has not converge at {} rounds with tol = {}'.format(i, tol))
                        pass
                if loop_break:
                    break
                else:
                    tol = tol * 5
            self.fast_ica_convergence = 'FastICA converges at {} rounds with tol = {}.'.format(i, tol)

        n_ch = sources.shape[0]
        n_t = sources.shape[1]
        artifacts = np.zeros((1, n_t))

        pool = Pool(11)
        multi_res = [pool.apply_async(PreProcessPipeline.get_artifact_wica, ([sources[i, :], n_t, wave_name, level, multipier],)) for i in range(n_ch)]
        pool.close()
        pool.join()
        for res in multi_res:
            temp = np.asarray(res.get()).reshape(1, -1)
            artifacts = np.concatenate((artifacts, temp), axis=0)

        self.tasks_cleaned._data[self.global_good_index] = self.tasks_cleaned.get_data(picks=self.global_good_name) - (np.dot(artifacts[1::, :].T, ica.mixing_.T) + ica.mean_).T
        del pool


    @staticmethod
    def get_artifact_wica(para):
        sources = para[0]
        n_t = para[1]
        wave_name = para[2]
        level = para[3]
        multipier = para[4]
        modulus = 2 ** level - n_t % 2 ** level
        sig = np.concatenate((sources, np.zeros(modulus))) if modulus != 0 else sources
        sig = matlab.double(sig.tolist())
        thresh, sorh, _ = PreProcessPipeline.engine.ddencmp('den', 'wv', sig, nargout=3)
        thresh = thresh * multipier
        swc = PreProcessPipeline.engine.swt(sig, level, wave_name)
        y = PreProcessPipeline.engine.wthresh(swc, sorh, thresh)
        w_ic = PreProcessPipeline.engine.iswt(y, wave_name)

        return w_ic[0][0:n_t]

    def mark_bad_epochs(self, drop_epoch=0.25, n_times=2, threshold=None, filtered_data=None, task_start=None,
                        task_end=None, trials_name=None, global_good_index=None, global_bads=None):
        self.start_point = []
        self.end_point = []
        self.epochs_point = []
        # tasks_values = list(trial_infomation.values())
        # np_tasks_values = np.array(tasks_values)
        # self.tasks_start = list(np_tasks_values[:, 0])
        # self.tasks_end = list(np_tasks_values[:, 1])
        # self.tasks_start = list(map(int, self.tasks_start))
        # self.tasks_end = list(map(int, self.tasks_end))
        epochs_info = OrderedDict()

        data = self.tasks_cleaned.get_data()
        for i in range(0, len(self.trial_duration), 2):
            task_data = np.zeros((self.n_ch, 1))
            trial_name = self.trial_name[int(i/2)]
            epochs_info[trial_name] = OrderedDict()

            start = int(self.trial_duration[i])
            end = int(self.trial_duration[i+1] - (self.trial_duration[i+1]-self.trial_duration[i]) % (n_times * self.tasks_cleaned.info['sfreq']))
            n_epochs = int((end - start) / (n_times * self.tasks_cleaned.info['sfreq']))

            data_cleaned_epochs = np.asarray(np.hsplit(data[self.global_good_index, start:end], n_epochs))
            data_epochs = np.asarray(np.hsplit(data[:, start:end], n_epochs))

            bad_channel_epochs = self.bad_epochs_faster(data_cleaned_epochs)

            for j in range(n_epochs):
                ch = self.get_ch_index(self.tasks_merged.ch_names, self.global_bads, bad_channel_epochs[j])
                ch_bad_index = np.argwhere(ch == 1)
                ch_bad_index = ch_bad_index.reshape(ch_bad_index.shape[0]).tolist()
                ch_bad_name = [self.tasks_merged.ch_names[i] for i in range(self.n_ch) if i in ch_bad_index]
                ratio = np.sum(ch) / len(ch)
                temp = mne.io.RawArray(data_epochs[j], info=self.tasks_merged.info.copy())
                temp.info['bads'] = ch_bad_name
                if ratio < drop_epoch:
                    temp = temp.interpolate_bads()
                    drop = 0
                    task_data = np.concatenate((task_data, temp.get_data()), axis=1)
                else:
                    drop = 1
                epochs_info[trial_name][str(j)] = {'epoch_data': temp.get_data().tolist(), 'bad_channel': ch_bad_name, 'interpolate_ratio': ratio, 'drop': drop}
            epochs_info[trial_name]['task_data'] = task_data[:, 1::].tolist()
        self.epochs_cleaned = epochs_info
        print(self.bad_channel_epochs)


    # def remove_bad_epochs(self, drop_epoch=0.25, n_times=2):
    #     epochs_info = OrderedDict()
    #     data = self.tasks_cleaned.get_data()
    #     for i in range(0, len(self.trial_duration), 2):
    #         task_data = np.zeros((self.n_ch, 1))
    #         trial_name = self.trial_name[int(i/2)]
    #         epochs_info[trial_name] = OrderedDict()
    #
    #         start = int(self.trial_duration[i])
    #         end = int(self.trial_duration[i+1] - (self.trial_duration[i+1]-self.trial_duration[i]) % (n_times * self.tasks_cleaned.info['sfreq']))
    #
    #         n_epochs = int((end - start) / (n_times * self.tasks_cleaned.info['sfreq']))
    #
    #         self.data_cleaned_epochs = np.asarray(np.hsplit(data[self.global_good_index, start:end], n_epochs))
    #         data_epochs = np.asarray(np.hsplit(data[:, start:end], n_epochs))
    #
    #         self.bad_channel_epochs = self.bad_epochs_faster(self.data_cleaned_epochs)
    #
    #         for j in range(n_epochs):
    #             ch = self.get_ch_index(self.tasks_merged.ch_names, self.global_bads, self.bad_channel_epochs[j])
    #             ch_bad_index = np.argwhere(ch == 1)
    #             ch_bad_index = ch_bad_index.reshape(ch_bad_index.shape[0]).tolist()
    #             ch_bad_name = [self.tasks_merged.ch_names[i] for i in range(self.n_ch) if i in ch_bad_index]
    #             ratio = np.sum(ch) / len(ch)
    #             temp = mne.io.RawArray(data_epochs[j], info=self.tasks_merged.info.copy())
    #             temp.info['bads'] = ch_bad_name
    #             if ratio < drop_epoch:
    #                 temp = temp.interpolate_bads()
    #                 drop = 0
    #                 task_data = np.concatenate((task_data, temp.get_data()), axis=1)
    #             else:
    #                 drop = 1
    #             epochs_info[trial_name][str(j)] = {'epoch_data': temp.get_data().tolist(), 'bad_channel': ch_bad_name, 'interpolate_ratio': ratio, 'drop': drop}
    #         epochs_info[trial_name]['task_data'] = task_data[:, 1::].tolist()
    #     self.epochs_cleaned = epochs_info


    def save_epochs_data(self, data_name=None, info_name=None, sheet_name=None):
        json.dump(self.epochs_cleaned, codecs.open(data_name, 'w', encoding='utf-8'), separators=(',', ':'), sort_keys=True)
        wb = load_workbook(info_name)
        sheet = wb.create_sheet(sheet_name)
        row = 1
        column = 1
        for task_name, task_data in self.epochs_cleaned.items():
            sheet.cell(row=row, column=column).value = task_name
            column += 1
            sheet.cell(row=row, column=column).value = "_".join(str(x) for x in self.global_bads)
            for data_type, data in task_data.items():
                if data_type != 'task_data':
                    bads = data['bad_channel']
                    bads.insert(0, data['drop'])
                    bads_str = "_".join(str(x) for x in bads)
                    column += 1
                    sheet.cell(row=row, column=column).value = bads_str

            row += 1
            column = 1

        sheet.cell(row=row+1, column=1).value = prep.fast_ica_convergence
        wb.save(info_name)
        wb.close()

    @staticmethod
    def read_epochs_data(fname=None):
        file_path = r'D:\EEGdata\clean_data_creativity\eeg_jan_29_2014'
        epochs_text = codecs.open(file_path, 'r', encoding='utf-8').read()
        epochs_data = json.loads(epochs_text)
        # print(123)

    @staticmethod
    def get_ch_index(ch_names, global_bads, local_bads):
        ch = np.ones(len(ch_names))
        good_ch_index = np.asarray([i for i in range(len(ch_names)) if ch_names[i] not in global_bads])
        good_ch_index = good_ch_index[np.argwhere(local_bads == 0)]
        ch[good_ch_index] = 0
        return ch

    @staticmethod
    def bad_epochs_faster(data, threshold):
        shape = data.shape
        n_epochs = shape[0]
        n_ch = shape[1]
        n_times = shape[2]
        criteria = []
        criteria.append(np.var(data, axis=2))
        criteria.append(np.median(np.gradient(data, axis=2), axis=2))
        criteria.append(np.ptp(data, axis=2))
        mean_epochs_channel = np.mean(data, axis=2)
        mean_epochs = np.mean(mean_epochs_channel, axis=1)
        criteria.append(mean_epochs_channel - mean_epochs.reshape(n_epochs, 1))
        res = np.zeros((n_epochs, n_ch))
        for i in range(len(criteria)):
            zscore = stats.zscore(criteria[i], axis=0)
            res += np.where(zscore > threshold[i], 1, 0)
        res = np.where(res > 0, 1, 0)
        return res

    def reference(self):
        pass

if __name__ == '__main__':
    clean_data_fname = r'D:\EEGdata\clean_data_six_problem\1_40\clean_data'
    clean_data_info = r'D:\EEGdata\clean_data_six_problem\clean_data_six_problem_1_40.xlsx'
    raw_data_fname = r'D:\EEGdata\raw_data_six_problem'
    subject_fname = r'D:\EEGdata\clean_data_six_problem\notebook.xlsx'
    subjects = read_subject_info(input_fname=r'D:\EEGdata\clean_data_six_problem\subjects.xlsx', sheet_name='subjects_1_40')
    tasks_combined = read_subject_info(r'D:\EEGdata\clean_data_six_problem\task_name_combined.xlsx', 'all')
    m_tasks = read_subject_info(r'D:\EEGdata\clean_data_six_problem\task_name_combined.xlsx', 'm_all')
    # raw_data_fname = r'D:\EEGdata\raw_data_creativity'
    # subject_fname = r'C:\Users\umroot\Desktop\creativity\notebook_res.xlsx'
    # subjects = read_subject_info(input_fname='C:\\Users\\umroot\\Desktop\\creativity\\subject.xlsx')
    for subject in subjects:
        print(subject)
        data_fname = raw_data_fname + "\\" + subject + "\\" +subject + ".vhdr"
        clean_data_save = clean_data_fname + "\\" +subject + ".json"
        raw = RawData()
        raw.read_raw_data(fname=data_fname, montage='D:\\workspace\\eeg_tool\\Cap63.locs', preload=True, scale=10e5)
        raw.read_trial_info(fname=subject_fname, sheet_name=subject)
        raw.split_tasks()
        prep = PreProcessPipeline(raw=raw.raw_data, tasks=raw.tasks_data, trial_info=raw.trial_info)
        prep.concatenate_tasks()
        res = prep.merge_task(tasks_combined)
        os.mkdir(r'D:\EEGdata\clean_data_six_problem\1_40\split_tasks' + "\\" + subject)
        for index, task_name in enumerate(tasks_combined):
            save_path = r'D:\EEGdata\clean_data_six_problem\1_40\split_tasks' + "\\" + subject +"\\" +m_tasks[index] +".mat"
            savemat(save_path, {'EEG':res[task_name]})

        # prep.filter(low_frequency=1., high_frequency=40.)
        # prep.remove_line_noise()
        # prep.remove_bad_channel()
        # prep.remove_artifact_wica()
        # prep.remove_bad_epochs()
        # prep.save_epochs_data(data_name=clean_data_save, info_name=clean_data_info, sheet_name=subject)

        # PreProcessPipeline.read_epochs_data()
