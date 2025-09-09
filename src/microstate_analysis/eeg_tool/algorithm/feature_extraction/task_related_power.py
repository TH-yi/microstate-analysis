#!/usr/bin/env python
# -*- coding: utf-8 -*-
# EEG signals respond differently to idea generation, idea evolution, and evaluation in a loosely controlled creativity experiment
# Time    : 2019-10-10
# Author  : Wenjun Jia
# File    : task_related_power.py


import matlab
import matlab.engine
import numpy as np
import codecs, json
from eeg_tool.utilis import read_subject_info
from scipy.signal import welch
from scipy.integrate import simps
from eeg_tool.algorithm.statistic.spss import electrodes_spss
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
from mne.time_frequency import psd_array_multitaper


class TaskRelatedPower:
    def __init__(self, task_data=None, band_frequency=[8,10,10,12], low_frequency=1, high_frequency=40):
        self.task_data = task_data
        self.band_frequency = band_frequency
        self.low_frequency = low_frequency
        self.high_frequency = high_frequency

    def band_powers(self, fs=500, nperseg=1000, noverlap=500, nfft=1000):
        res = {}
        for task_name, task_data in self.task_data.items():
            res[task_name] = {}
            for epoch_name, epoch_data in task_data.items():
                if epoch_name == 'task_data':
                    sig = np.asarray(epoch_data)
                    if sig.shape[1] == 0:
                        continue
                elif epoch_data['drop'] == 0:
                    sig = np.asarray(epoch_data['epoch_data'])
                else:
                    continue
                # power by Welch
                freq, pxx = welch(sig, fs=fs, nperseg=nperseg, noverlap=noverlap, nfft=nfft)
                total_power = TaskRelatedPower.band_power(freq, pxx, self.low_frequency, self.high_frequency)

                # power by fft
                # fft_values = np.power(np.absolute(np.fft.rfft(sig, axis=1)), 2)
                # fft_freq = np.fft.rfftfreq(sig.shape[1], 1./fs)

                # power by multitaper
                # pxx, freq = psd_array_multitaper(sig, fs, fmin=self.low_frequency, fmax=self.high_frequency, adaptive=True, normalization='full', verbose=0)
                # total_power = TaskRelatedPower.band_power(freq, pxx, self.low_frequency, self.high_frequency)
                # total_power = TaskRelatedPower.band_power_fft(fft_freq, fft_values, self.low_frequency, self.high_frequency)
                res[task_name][epoch_name] = {'total_power': total_power.tolist(), 'total_log_power': (np.log10(total_power)).tolist()}
                for i in range(0, len(self.band_frequency), 2):
                    # power = TaskRelatedPower.band_power(freq, pxx, self.band_frequency[i], self.band_frequency[i + 1])
                    # power = TaskRelatedPower.band_power_fft(fft_freq, fft_values, self.band_frequency[i], self.band_frequency[i+1])
                    power = TaskRelatedPower.band_power(freq, pxx, self.band_frequency[i], self.band_frequency[i + 1])
                    freq_str = str(self.band_frequency[i]) + "_" + str(self.band_frequency[i+1])
                    res[task_name][epoch_name][freq_str] = {'band_power': power.tolist(), 'band_log_power': (np.log10(power)).tolist(), 'band_relative_power': (abs(power/total_power)).tolist()}
        return res

    def task_related_power(self, band_powers, active_task, reference_task, task_data_name='task_data', power_name='band_log_power'):
        electrodes_position = electrodes_spss()
        n_ch = len(electrodes_position)
        res = {}
        spss = {}
        for k in range(len(active_task)):
            name = active_task[k].split('_')[1]
            if name not in res:
                res[name] = {}
            for i in range(0, len(self.band_frequency), 2):
                band = str(self.band_frequency[i]) + "_" + str(self.band_frequency[i + 1])
                trp = np.asarray(band_powers[active_task[k]][task_data_name][band][power_name]) - np.asarray(band_powers[reference_task[k]][task_data_name][band][power_name])
                trp_spss = trp[electrodes_position].reshape(n_ch, -1)
                if band not in res[name]:
                    res[name][band] = trp_spss
                else:
                    res[name][band] = np.concatenate((res[name][band], trp_spss), axis=1)
        for task_name, task_data in res.items():
            spss[task_name] = {}
            for band_name, band_data in task_data.items():
                spss[task_name][band_name] = np.mean(band_data, axis=1)
                # spss[task_name][band_name] = band_data

        return spss


    def write_trp_excel(self, fname, data, task_name):
        wb = load_workbook(fname)
        for i in range(0, len(self.band_frequency), 2):
            sheet_name = str(self.band_frequency[i]) + "_" + str(self.band_frequency[i+1])
            sheet = wb.create_sheet(sheet_name)
            row = 1
            for subject_name, subject_data in data.items():
                col = 1
                for name in task_name:
                    row_data = subject_data[name][sheet_name]
                    for k in range(row_data.shape[0]):
                        if row_data.ndim == 2:
                            for k1 in range(row_data.shape[1]):
                                sheet.cell(row, col).value = row_data[k, k1]
                                col += 1
                        else:
                            sheet.cell(row, col).value = row_data[k]
                            col += 1
                row += 1
        wb.save(fname)
        wb.close()


    @staticmethod
    def band_power(freq, pxx, low, high):
        n_ch = pxx.shape[0]
        idx = np.logical_and(freq >= low, freq <= high).reshape(1, -1)
        idx = np.repeat(idx, pxx.shape[0], axis=0)
        power = simps(pxx[idx].reshape(n_ch, -1), dx=freq[1] - freq[0], axis=1)
        return power

    @staticmethod
    def band_power_fft(fft_freq, fft_values, low, high):
        n_ch = fft_values.shape[0]
        idx = np.logical_and(fft_freq >= low, fft_freq <= high).reshape(1, -1)
        idx = np.repeat(idx, n_ch, axis=0)
        # total_power = np.sum(fft_values[idx].reshape(n_ch, -1), axis=1)
        return np.mean(fft_values[idx].reshape(n_ch, -1), axis=1)


if __name__ == '__main__':
    clean_data_fname = r'D:\EEGdata\clean_data_creativity\1_50_2s_multipaer'
    # subjects = read_subject_info(input_fname='C:\\Users\\umroot\\Desktop\\creativity\\subject.xlsx')
    subjects = read_subject_info(r'D:\EEGdata\clean_data_creativity\clean_subjects_28.xlsx')

    trp_excel = r'D:\EEGdata\clean_data_creativity\1_50_2s_multipaer\trp_28_8_10_10_12hz.xlsx'
    trp_res = OrderedDict()
    # task_name = ['idea generation', 'idea evolution', 'idea rating']
    # active_task = ['1_idea evolution', '2_idea evolution', '3_idea evolution']
    # reference_task = ['1_rest', '1_rest', '1_rest']
    active_task = []
    reference_task = []

    task_name = ['idea generation', 'idea evolution', 'idea rating']

    # task_name = ['read problem', 'generate solution', 'rate generation', 'evaluate solution', 'type', 'rate evaluation']


    for i in range(1, 4):
        for j in range(len(task_name)):
            active_task.append(str(i)+"_"+task_name[j])
            reference_task.append('1_rest')

    # task_name = ['1_idea generation', '2_idea generation', '3_idea generation',
    #              '1_idea evolution', '2_idea evolution', '3_idea evolution',
    #              '1_idea rating', '2_idea rating', '3_idea rating']
    #
    # rest_name = ['1_rest']

    for subject in subjects:
        print(subject)
        # data_fname_save = clean_data_fname +"\\" + subject +"_epochs_power" +".json"
        data_fname = clean_data_fname + "\\" +subject + "_power" + ".json"
        # data_fname = r'D:\EEGdata\clean_data_six_problem\task_related_power' + "\\" +subject + "_power" + ".json"

        # data_fname = clean_data_fname + "\\" +subject + ".json"

        data_text = epochs_text = codecs.open(data_fname, 'r', encoding='utf-8').read()
        data = json.loads(data_text)
        # power = TaskRelatedPower(data)

        # res = power.band_powers()
        power = TaskRelatedPower()
        # res = {'task': power.concatenate_tasks_powers(data, task_name), 'rest': power.concatenate_tasks_powers(data, ['1_rest'])}
        # json.dump(res, codecs.open(data_fname_save, 'w', encoding='utf-8'), separators=(',', ':'), sort_keys=True, indent=4)

        trp_res[subject] = power.task_related_power(data, active_task, reference_task)
    power = TaskRelatedPower()
    power.write_trp_excel(trp_excel, trp_res, task_name)
