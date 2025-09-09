import pandas as pd
import os
from openpyxl import Workbook, load_workbook
import collections
import numpy as np
import mne
import re
import codecs, json
import os

def get_root_path():
    return os.path.abspath(os.path.join(os.getcwd(), "../../.."))

def read_info(fname, sname):
    wb = load_workbook(fname)
    sheet = wb.get_sheet_by_name(sname)
    res = []
    for i, item in enumerate(sheet.rows):
        sub = []
        for j, j_item in enumerate(item):
            sub.append(sheet.cell(row=i + 1, column=j + 1).value)
        res.append(sub)
    return res

def write_append_info(path_dir, file_path, data):
    fname = os.path.basename(file_path)
    if fname in os.listdir(path_dir):
        data_save = json.loads(codecs.open(file_path, 'r', encoding='utf-8').read())
        data_save.append(data)
    else:
        data_save = [data]
    json.dump(data_save, codecs.open(file_path, 'w', encoding='utf-8'),separators=(',', ':'), sort_keys=True)


def write_info(fname, sname, data, create_workbook=False, row_first=True):
    wb = Workbook() if create_workbook else load_workbook(fname)
    sheet = wb.create_sheet(sname)
    sheet.title = sname
    for i in range(len(data)):
        for j in range(len(data[i])):
            if row_first:
                sheet.cell(row=i + 1, column=j + 1).value = data[i][j]
            else:
                sheet.cell(row=j + 1, column=i + 1).value = data[i][j]
    wb.save(fname)
    wb.close()

def write_trial_info(input_fname, trial_info):
    wb = load_workbook(input_fname)
    for sheet_name, trial in trial_info.items():
        sheet = wb.get_sheet_by_name(sheet_name)
        for i in range(2, 13):
            start = sheet.cell(row=i, column=5).value
            end = sheet.cell(row=i, column=6).value
            if ',' in str(start):
                start = start.split(',')[0]
                end = end.split(',')[-1]
            start = int(trial[str(start)])
            end = int(trial[str(end)])
            sheet.cell(row=i, column=3).value = start
            sheet.cell(row=i, column=4).value = end
    wb.save(input_fname)
    wb.close()


def read_trial_info(fname, sheet_name):
    wb = load_workbook(fname)
    sheet = wb.get_sheet_by_name(sheet_name)
    res = {}
    for i, item in enumerate(sheet.rows):
        flag = True
        for j, j_item in enumerate(item):
            value = sheet.cell(row=i + 1, column=j + 1).value
            if value and value != sheet_name:
                if flag:
                    key = value
                    res[key] = []
                    flag = False
                else:
                    res[key].append(value)
    del res['Name']
    wb.close()
    return res


def read_subject_info(input_fname, sheet_name):
    sub = []
    wb = load_workbook(input_fname)
    sheet = wb.get_sheet_by_name(sheet_name)
    for i, item in enumerate(sheet.rows):
        for j, j_item in enumerate(item):
            if sheet.cell(row=i + 1, column=j + 1).value:
                sub.append(sheet.cell(row=i + 1, column=j + 1).value)
    wb.close()
    return sub

def read_xlsx(input_fname, sheet_name, row=True):
    sub = []
    wb = load_workbook(input_fname)
    sheet = wb.get_sheet_by_name(sheet_name)
    for i, item in enumerate(sheet.rows):
        temp = []
        for j, j_item in enumerate(item):
            if sheet.cell(row=i + 1, column=j + 1).value:
                temp.append(sheet.cell(row=i + 1, column=j + 1).value)
        sub.append(temp)
    wb.close()
    return sub if row else np.asarray(sub).T.tolist()


def read_vmrk_write_info(subjects, base_path=r'D:\EEGdata\raw_data_six_problem', notebook_path=r'D:\EEGdata\raw_data_six_problem\record_notebook\cleaned_subject_notebook.xlsx'):
    wb = load_workbook(notebook_path)
    for subject in subjects:
        path = base_path + "\\" + subject + "\\" + subject +".vmrk"
        sheet = wb.get_sheet_by_name(subject)
        info = []
        for i in range(1, sheet.max_row):
            temp = []
            start = str(sheet.cell(row=i+1, column=5).value)
            end = str(sheet.cell(row=i+1, column=6).value)
            if ',' in start:
                start_temp = start.split(',')
                end_temp = end.split(',')
                for j in range(len(start_temp)):
                    temp.append(start_temp[j])
                    temp.append(end_temp[j])
            else:
                temp.append(start)
                temp.append(end)
            info.append(temp)

        with open(path, 'r') as f:
            lines = f.readlines()

        for row_index, row in enumerate(info):
            start = []
            end = []
            for i in range(0, len(row), 2):
                for line in lines:
                    if '=Stimulus' in line:
                        line_temp = line.split(',')
                        index = line_temp[0].split('=')[0].split('Mk')[1]
                        index_value = str(line_temp[2])
                        if index == row[i]:
                            start.append(index_value)
                        elif index == row[i+1]:
                            end.append(index_value)
            sheet.cell(row=row_index + 2, column=3).value = ','.join(start)
            sheet.cell(row=row_index + 2, column=4).value = ','.join(end)

    wb.save(notebook_path)
    wb.close()




def read_data(path):
    with open(path, 'r') as f:
       lines = f.readlines()
    matrix = []
    for line in lines:
        res = []
        temp = line.split(' ')
        for num in temp:
            if num:
                res.append(float(num))
        matrix.append(res)
    return np.asarray(matrix)



def create_info(sfreq):
    ch = ['Fp1', 'Fz', 'F3', 'F7', 'FT9', 'FC5', 'FC1', 'C3', 'T7', 'TP9', 'CP5', 'CP1', 'Pz', 'P3', 'P7', 'O1', 'Oz', 'O2', 'P4', 'P8', 'TP10', 'CP6', 'CP2', 'C4', 'T8', 'FT10', 'FC6', 'FC2', 'F4', 'F8', 'Fp2', 'AF7', 'AF3', 'AFz', 'F1', 'F5', 'FT7', 'FC3', 'FCz', 'C1', 'C5', 'TP7', 'CP3', 'P1', 'P5', 'PO7', 'PO3', 'POz', 'PO4', 'PO8', 'P6', 'P2', 'CPz', 'CP4', 'TP8', 'C6', 'C2', 'FC4', 'FT8', 'F6', 'F2', 'AF4', 'AF8']
    montage = 'D:\\workspace\\eeg_tool\\Cap63.locs'
    info = mne.create_info(ch_names=ch, sfreq=sfreq, ch_types='eeg', montage=os.path.join(get_root_path(),'eeg_tool','Cap63.locs'))
    return info

def load_data(path):
    data_text = codecs.open(path, 'r', encoding='utf-8').read()
    data = json.loads(data_text)
    return data

def to_string(data):
    res = [str(i) for i in data]
    return '-'.join(res)


def load_config_design():
    subjects = read_subject_info(r'D:\EEGdata\clean_data_six_problem\subjects.xlsx', 'subjects_1_40')
    tasks = read_subject_info(r'D:\EEGdata\clean_data_six_problem\task_name_combined.xlsx', 'm_all_ordered')
    conditions = read_subject_info(r'D:\EEGdata\clean_data_six_problem\condition_name.xlsx', 'm_all_ordered')
    res = {}
    for condition in conditions:
        res[condition] = []
        for task in tasks:
            if task.startswith(condition):
                res[condition].append(task)
    return subjects, tasks, conditions, res

if __name__ == '__main__':
    # res = read_trial_info('C:\\Users\\umroot\\Desktop\\six_problem\\notebook_res.xlsx', 'april_02(3)')
    # for key, value in res.items():
    #     print(key,value)
    subjects = read_subject_info(r'D:\EEGdata\raw_data_six_problem\record_notebook\cleaned_subjects.xlsx')
    # subjects = ['april_19(2)']
    read_vmrk_write_info(subjects)